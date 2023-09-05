import openpyxl
from ping3 import ping, verbose_ping

def ping_hosts(input_file, output_file):
    try:
        # Charger le fichier Excel d'entrée
        wb_input = openpyxl.load_workbook(input_file)
        sheet_input = wb_input.active

        # Créer un nouveau fichier Excel pour les résultats
        wb_output = openpyxl.Workbook()
        sheet_output = wb_output.active

        # Entête du fichier de sortie
        sheet_output.append(["Host", "IP Address", "Status"])

        # Itérer sur les lignes du fichier d'entrée
        for row in sheet_input.iter_rows(min_row=2, values_only=True):
            host = row[0]
            ip_address = None

            # Effectuer le ping de l'hôte
            try:
                ip_address = verbose_ping(host, count=1, timeout=10)
                status = "UP" if ip_address else "DOWN"
            except:
                status = "DOWN"

            # Ajouter les résultats au fichier de sortie
            sheet_output.append([host, ip_address, status])

        # Enregistrer le fichier de sortie
        wb_output.save(output_file)
        print("Ping terminé avec succès. Les résultats sont enregistrés dans", output_file)
    except Exception as e:
        print("Une erreur s'est produite:", e)

if __name__ == "__main__":
    input_excel_file = r"C:\Users\inputFile.xlsx"
    output_excel_file = r"C:\Users\output.xlsx"
    ping_hosts(input_excel_file, output_excel_file)
